from pathlib import Path
import openpyxl
from src.controllers.manager import Manager
from src.models.base.application import aplicacion
from src.strategies.k_qnodes import KQNodes


_docs = Path.home() / "Documents"
RUTA_EXCEL = None
for _d in _docs.iterdir():
    if _d.is_dir() and "analisis" in _d.name.lower():
        _candidato = _d / "DatosPruebas2026_1.xlsx"
        if _candidato.exists():
            RUTA_EXCEL = _candidato
            break
if RUTA_EXCEL is None:
    RUTA_EXCEL = _docs / "Analisis y diseno de algoritmos" / "DatosPruebas2026_1.xlsx"




def obtener_hoja(n_nodos: int, pagina: str) -> str | None:
    if not RUTA_EXCEL or not RUTA_EXCEL.exists():
        return None
    import re
    wb = openpyxl.load_workbook(RUTA_EXCEL)
    patron = re.compile(rf"^{n_nodos}{pagina}-Elementos\s*$")
    for name in wb.sheetnames:
        if patron.match(name):
            wb.close()
            return name
    wb.close()
    return None




COLUMNAS_QNODES = {
    2: ("D", "E", "F"),
    3: ("J", "K", "L"),
    4: ("P", "Q", "R"),
    5: ("V", "W", "X"),
}


ABECEDARIO = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"




def binario_a_letras(bits: str) -> str:
    return "".join(ABECEDARIO[i] for i, b in enumerate(bits) if b == "1")




def buscar_fila(ws, alcance_letras: str, mecanismo_letras: str) -> int:
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=3, values_only=False):
        b_val = str(row[1].value).strip() if row[1].value else ""
        c_val = str(row[2].value).strip() if row[2].value else ""
        if b_val == alcance_letras and c_val == mecanismo_letras:
            return row[0].row
    return None




def escribir_resultados(ruta: Path, hoja: str, fila: int, k: int, particion: str, perdida: float, tiempo: float):
    col_part, col_perd, col_tmp = COLUMNAS_QNODES[k]
    try:
        wb = openpyxl.load_workbook(ruta)
        ws = wb[hoja]
        ws[f"{col_part}{fila}"] = particion
        ws[f"{col_perd}{fila}"] = round(perdida, 6)
        ws[f"{col_tmp}{fila}"] = f"Segundos: {tiempo:.4f}"
        wb.save(ruta)
        wb.close()
        print(f"  -> Escrito en {col_part}{fila}, {col_perd}{fila}, {col_tmp}{fila}")
    except PermissionError:
        print(f"  !! ERROR: No se pudo escribir en el Excel. Cerrá el archivo y volvé a ejecutar.")
    except Exception as e:
        print(f"  !! ERROR al escribir Excel: {e}")




def iniciar():
     
    estado_inicial = "10000000000000000000"
    condiciones =    "11111111111111111111"
    alcance =        "11111111111111111111"
    #                 ABCDEFGHIJKLMNOPQRST
    mecanismo =      "01111111111111111110"







    n_nodos = len(estado_inicial)
    pagina = aplicacion.pagina_red_muestra
    hoja = obtener_hoja(n_nodos, pagina)


    alcance_letras = binario_a_letras(alcance)
    mecanismo_letras = binario_a_letras(mecanismo)


    gestor_redes = Manager(estado_inicial)

    gestor_redes.generar_red(
        dimensiones=len(estado_inicial),
        datos_deterministas=True
    )

    print("Archivo TPM:", gestor_redes.tpm_filename)
    mpt = gestor_redes.cargar_red()


    fila = None
    if hoja and RUTA_EXCEL and RUTA_EXCEL.exists():
        print(f"Hoja Excel: {hoja}")
        wb = openpyxl.load_workbook(RUTA_EXCEL)
        ws = wb[hoja]
        fila = buscar_fila(ws, alcance_letras, mecanismo_letras)
        wb.close()


    if fila:
        print(f"Fila encontrada: {fila}  (alcance={alcance_letras}, mecanismo={mecanismo_letras})")
    else:
        print(f"No se encontró fila para alcance={alcance_letras}, mecanismo={mecanismo_letras}")
        print("Se ejecutará igual pero sin guardar en Excel.")
        fila = None


    k_valores = [3,4,5]
    for k in k_valores:
        print(f"\n--- KQNodes k={k} ---")
        analizador = KQNodes(mpt, k=k)
        sol = analizador.aplicar_estrategia(estado_inicial, condiciones, alcance, mecanismo)
        print(sol)
        if fila:
            escribir_resultados(RUTA_EXCEL, hoja, fila, k, sol.particion, sol.perdida, sol.tiempo_ejecucion)


